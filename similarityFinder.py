from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles import Font
from openpyxl.workbook import Workbook
def main():
    fileName = input("Enter the filepath of your Excel Worksheet: ")
    wb = load_workbook(filename=fileName)

    in_sheet_name = input("Enter the name of the input worksheet: ")
    input_sheet = wb[in_sheet_name]
    out_sheet_name = input("Enter the name of the output worksheet: ")
    output_sheet = wb[out_sheet_name]
    max_output_row = output_sheet.max_row
    max_input_row = input_sheet.max_row

    #Only for matching one column right now
    in_col = input("Enter the letter of the column you want to copy data from: ")
    in_col = in_col + '4' # --------- Do I need to do this?? -----------
    # getting column number from name : https://stackoverflow.com/a/12902801
    xy = coordinate_from_string(in_col) # returns ('A',4)
    in_col = column_index_from_string(xy[0])
    out_col = input("Enter the letter of the column you want to copy data to: ") 
    out_col = out_col + '4'
    xy = coordinate_from_string(out_col) 
    out_col = column_index_from_string(xy[0]) 
    
    for i in range(1,max_output_row+1):
        cell_out = output_sheet.cell(row = i, column = 2)
        exists = False
        for j in range(1,max_input_row+1):
            cell_in = input_sheet.cell(row = j, column = 2)
            if conditions(cell_in,cell_out):
                output_sheet.cell(row=i,column=out_col).value = '=\'CTB 20\'!AM{}'.format(j)
                exists = True
        if not exists and cell_out.value != None:
            cell = output_sheet.cell(row=i, column=out_col)
            cell.value = 'CHECK'
            cell.font = Font(color='FFFF0000',bold=True)
    wb.save("outputfile.xlsx")   
           
def conditions(cell_in,cell_out):
    if cell_out.value == cell_in.value:
        if cell_in.value != None and "Other" not in cell_in.value and "other" not in cell_in.value:
            return True
    return False

main()